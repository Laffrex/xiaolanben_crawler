import os
import time
import pandas as pd
from openpyxl import load_workbook
import shutil

class ExcelFileManager:
    """Excel文件管理器，负责表格的集中初始化与验证"""
    
    # 类级别锁，确保同一时间只有一个实例在操作文件
    _file_locks = {}
    
    def __init__(self, file_path):
        """初始化Excel文件管理器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        # 定义所有支持的表格结构
        self.table_structures = {
            'APP': ['产品名', '产品链接'],
            'Website': ['网站名', '网站链接'],
            '微信公众号': ['微信公众号', '链接'],
            '微信小程序': ['微信小程序', '链接'],
            '其他媒体': ['其他媒体', '链接'],
            '集团成员': ['成员名', '成员链接'],
            '对外投资': ['被投资方', '被投资方链接'],
            '投资方': ['投资方', '投资方链接']
        }
    
    @classmethod
    def _acquire_lock(cls, file_path):
        """获取文件锁"""
        start_time = time.time()
        timeout = 30  # 30秒超时
        while file_path in cls._file_locks and cls._file_locks[file_path]:
            if time.time() - start_time > timeout:
                raise TimeoutError(f"获取文件锁超时: {file_path}")
            print(f"等待文件锁释放: {file_path}")
            time.sleep(0.5)
        
        cls._file_locks[file_path] = True
        print(f"获取文件锁: {file_path}")
    
    @classmethod
    def _release_lock(cls, file_path):
        """释放文件锁"""
        if file_path in cls._file_locks:
            cls._file_locks[file_path] = False
            print(f"释放文件锁: {file_path}")
    
    def initialize_tables(self, required_tables=None):
        """初始化Excel文件的表格结构
        
        Args:
            required_tables: 指定需要初始化的表格，默认为None表示初始化所有表格
            
        Returns:
            bool: 初始化是否成功
        """
        # 如果未指定需要的表格，则使用所有支持的表格
        if required_tables is None:
            tables_to_init = self.table_structures
        else:
            # 只初始化指定的表格
            tables_to_init = {name: self.table_structures[name] 
                             for name in required_tables 
                             if name in self.table_structures}
        
        max_retries = 3
        for retry in range(max_retries):
            try:
                # 获取文件锁
                self._acquire_lock(self.file_path)
                
                print(f"初始化Excel文件: {self.file_path}，尝试次数: {retry+1}/{max_retries}")
                
                # 检查文件是否存在
                if os.path.exists(self.file_path):
                    # 验证文件是否可访问
                    try:
                        # 尝试加载现有文件
                        book = load_workbook(self.file_path)
                        
                        # 检查每个表格是否存在，如果不存在则创建
                        missing_sheets = [sheet_name for sheet_name in tables_to_init 
                                         if sheet_name not in book.sheetnames]
                        
                        if missing_sheets:
                            print(f"需要创建以下表格: {', '.join(missing_sheets)}")
                            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                                writer.book = book
                                
                                for sheet_name in missing_sheets:
                                    # 创建空的DataFrame并保存
                                    df = pd.DataFrame(columns=tables_to_init[sheet_name])
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                                    print(f"创建表格: {sheet_name}")
                        else:
                            print("所有必要的表格已存在，无需创建")
                            
                    except Exception as e:
                        print(f"读取现有文件时出错: {str(e)}")
                        # 文件可能已损坏，创建备份并重新创建
                        backup_file = f"{self.file_path}.bak"
                        try:
                            shutil.copy2(self.file_path, backup_file)
                            print(f"已备份可能损坏的文件: {backup_file}")
                        except:
                            print("无法备份文件")
                            
                        os.remove(self.file_path)
                        print(f"删除损坏的文件: {self.file_path}")
                        raise Exception("文件已损坏，将重新创建")
                
                # 如果文件不存在或已删除，创建新文件
                if not os.path.exists(self.file_path):
                    with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                        for sheet_name, columns in tables_to_init.items():
                            # 创建空的DataFrame并保存
                            df = pd.DataFrame(columns=columns)
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            print(f"创建表格: {sheet_name}")
                
                # 初始化完成后进行验证
                if self._verify_tables(tables_to_init):
                    print("表格初始化并验证成功")
                    return True
                else:
                    raise Exception("表格验证失败")
                
            except Exception as e:
                print(f"初始化表格时出错: {str(e)}")
                if retry < max_retries - 1:
                    print(f"准备第 {retry+2} 次尝试...")
                    time.sleep(2)  # 等待一下再重试
                else:
                    print(f"已达到最大重试次数 ({max_retries})，初始化失败")
                    # 尝试备用方案，如创建临时文件
                    return self._fallback_initialization(tables_to_init)
            finally:
                # 释放文件锁
                self._release_lock(self.file_path)
        
        return False
    
    def _verify_tables(self, expected_tables):
        """验证表格是否成功初始化
        
        Args:
            expected_tables: 期望的表格结构字典
            
        Returns:
            bool: 验证是否通过
        """
        try:
            if not os.path.exists(self.file_path):
                print(f"文件不存在: {self.file_path}")
                return False
            
            # 尝试打开文件并验证每个表格
            excel_file = pd.ExcelFile(self.file_path)
            for sheet_name, expected_columns in expected_tables.items():
                if sheet_name not in excel_file.sheet_names:
                    print(f"缺少表格: {sheet_name}")
                    return False
                
                # 验证列名
                df = pd.read_excel(self.file_path, sheet_name=sheet_name)
                if not all(col in df.columns for col in expected_columns):
                    print(f"表格 {sheet_name} 的列名不匹配")
                    return False
            
            return True
            
        except Exception as e:
            print(f"验证表格时出错: {str(e)}")
            return False
    
    def _fallback_initialization(self, tables_to_init):
        """初始化失败时的备用方案
        
        Args:
            tables_to_init: 需要初始化的表格结构
            
        Returns:
            bool: 备用初始化是否成功
        """
        try:
            print("尝试备用初始化方案...")
            # 生成临时文件名
            temp_file = f"{os.path.splitext(self.file_path)[0]}_temp.xlsx"
            print(f"创建临时文件: {temp_file}")
            
            # 创建临时文件
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                for sheet_name, columns in tables_to_init.items():
                    df = pd.DataFrame(columns=columns)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 验证临时文件
            if os.path.exists(temp_file) and self._verify_file(temp_file, tables_to_init):
                # 用临时文件替换原文件
                shutil.move(temp_file, self.file_path)
                print(f"已用临时文件替换原文件: {self.file_path}")
                return True
            else:
                print("临时文件创建或验证失败")
                return False
                
        except Exception as e:
            print(f"备用初始化方案失败: {str(e)}")
            return False
    
    def _verify_file(self, file_path, expected_tables):
        """验证指定文件的表格结构
        
        Args:
            file_path: 要验证的文件路径
            expected_tables: 期望的表格结构
            
        Returns:
            bool: 验证是否通过
        """
        try:
            excel_file = pd.ExcelFile(file_path)
            for sheet_name in expected_tables:
                if sheet_name not in excel_file.sheet_names:
                    return False
            return True
        except:
            return False 